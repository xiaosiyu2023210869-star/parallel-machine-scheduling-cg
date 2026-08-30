"""WSPT-based heuristic baseline for saved CG comparison instances.

The baseline is intentionally solver-free.  It reads the same fixed instances
used by the comparison experiment, builds a WSPT/LRF schedule, then improves
each machine sequence by reordering tool-life blocks with a batch-level WSPT
index that includes the tool-change time.
"""

from __future__ import annotations

import csv
import json
import math
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .parameters import effective_machine_count, machine_count_adjusted


DEFAULT_SCALE_PARAMS = {
    "small": {"m": 3, "c": 4, "theta": 2.0},
    "medium": {"m": 3, "c": 2, "theta": 4.0},
    "large": {"m": 3, "c": 2, "theta": 4.0},
}

METHOD_NAME = "WSPT-BR"
STATUS = "FEASIBLE"


@dataclass(frozen=True)
class InstanceRecord:
    scale: str
    instance_id: int
    n: int
    p: list[float]
    w: list[float]


@dataclass(frozen=True)
class HeuristicResult:
    objective: float
    time_sec: float
    schedule: list[list[int]]
    wspt_lrf_objective: float
    batch_reorder_improvement_pct: float


def wspt_key(job: int, p: list[float], w: list[float]) -> tuple[float, float, float, int]:
    return (p[job] / max(w[job], 1e-12), p[job], -w[job], job)


def compute_sequence_cost(seq: Iterable[int], p: list[float], w: list[float], c: int, theta: float) -> float:
    cur = 0.0
    cost = 0.0
    for pos, job in enumerate(seq):
        if pos > 0 and pos % c == 0:
            cur += theta
        cur += p[job]
        cost += w[job] * cur
    return float(cost)


def schedule_objective(machine_seq: list[list[int]], p: list[float], w: list[float], c: int, theta: float) -> float:
    return float(sum(compute_sequence_cost(seq, p, w, c, theta) for seq in machine_seq if seq))


def validate_schedule(machine_seq: list[list[int]], n: int) -> None:
    assigned = [job for seq in machine_seq for job in seq]
    if len(assigned) != n or sorted(assigned) != list(range(n)):
        raise ValueError("Heuristic produced an invalid schedule.")


def wspt_lrf_schedule(p: list[float], w: list[float], m: int) -> list[list[int]]:
    """Largest-ratio-first assignment with WSPT order and least current load."""

    machine_seq: list[list[int]] = [[] for _ in range(m)]
    loads = [0.0] * m
    for job in sorted(range(len(p)), key=lambda j: wspt_key(j, p, w)):
        machine = min(range(m), key=lambda i: (loads[i], len(machine_seq[i]), i))
        machine_seq[machine].append(job)
        loads[machine] += p[job]
    return machine_seq


def block_reordered_sequence(seq: list[int], p: list[float], w: list[float], c: int, theta: float) -> list[int]:
    """Sort jobs inside each block by WSPT, then sort blocks by (P+theta)/W."""

    blocks = []
    for block_index, start in enumerate(range(0, len(seq), c)):
        block = sorted(seq[start : start + c], key=lambda j: wspt_key(j, p, w))
        block_p = sum(p[j] for j in block)
        block_w = sum(w[j] for j in block)
        block_ratio = (block_p + theta) / max(block_w, 1e-12)
        blocks.append((block_ratio, block_p, -block_w, block_index, block))

    reordered = []
    for *_prefix, block in sorted(blocks):
        reordered.extend(block)
    return reordered


def improved_wspt_schedule(p: list[float], w: list[float], m: int, c: int, theta: float) -> list[list[int]]:
    base = wspt_lrf_schedule(p, w, m)
    improved = [block_reordered_sequence(seq, p, w, c, theta) for seq in base]
    validate_schedule(improved, len(p))
    return improved


def run_improved_wspt(p: list[float], w: list[float], m: int, c: int, theta: float) -> HeuristicResult:
    m = effective_machine_count(m, len(p))
    if c <= 0:
        raise ValueError("c must be positive.")

    start = time.perf_counter()
    base_schedule = wspt_lrf_schedule(p, w, m)
    base_obj = schedule_objective(base_schedule, p, w, c, theta)
    schedule = [block_reordered_sequence(seq, p, w, c, theta) for seq in base_schedule]
    validate_schedule(schedule, len(p))
    objective = schedule_objective(schedule, p, w, c, theta)
    elapsed = time.perf_counter() - start
    improvement = 0.0 if base_obj == 0 else 100.0 * (base_obj - objective) / base_obj
    return HeuristicResult(
        objective=objective,
        time_sec=elapsed,
        schedule=schedule,
        wspt_lrf_objective=base_obj,
        batch_reorder_improvement_pct=improvement,
    )


def load_instances(data_dir: Path) -> dict[int, InstanceRecord]:
    index: dict[int, InstanceRecord] = {}
    for scale in ("small", "medium", "large"):
        path = data_dir / f"instances_{scale}.json"
        records = json.loads(path.read_text(encoding="utf-8"))
        for item in records:
            inst = InstanceRecord(
                scale=scale,
                instance_id=int(item["id"]),
                n=int(item["n"]),
                p=[float(value) for value in item["p"]],
                w=[float(value) for value in item["w"]],
            )
            index[inst.instance_id] = inst
    return index


def read_table(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader.fieldnames or []), list(reader)

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(value) if value is not None else "" for value in rows[0]]
    data = []
    for values in rows[1:]:
        data.append({key: values[index] if index < len(values) else None for index, key in enumerate(header)})
    return header, data


def parse_int(value: object, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    return int(float(value))


def parse_float(value: object, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(result):
        return default
    return result


def build_wspt_rows(
    source_rows: list[dict[str, object]],
    instances: dict[int, InstanceRecord],
    *,
    machines: int | None = None,
) -> list[dict[str, object]]:
    cg_rows = [row for row in source_rows if str(row.get("method")) == "CG"]
    wspt_rows: list[dict[str, object]] = []
    for row in cg_rows:
        instance_id = parse_int(row.get("instance_id"))
        if instance_id is None or instance_id not in instances:
            continue
        inst = instances[instance_id]
        defaults = DEFAULT_SCALE_PARAMS[inst.scale]
        requested_m = machines or parse_int(row.get("m"), int(defaults["m"])) or int(defaults["m"])
        m = effective_machine_count(requested_m, inst.n)
        c = parse_int(row.get("c"), int(defaults["c"])) or int(defaults["c"])
        theta = parse_float(row.get("theta"), float(defaults["theta"])) or float(defaults["theta"])

        result = run_improved_wspt(inst.p, inst.w, m, c, theta)
        cg_obj = parse_float(row.get("objective"))
        gap_vs_cg = None if cg_obj in (None, 0.0) else 100.0 * (result.objective - cg_obj) / cg_obj
        wspt_rows.append(
            {
                "instance_id": instance_id,
                "n": inst.n,
                "method": METHOD_NAME,
                "objective": result.objective,
                "time_sec": result.time_sec,
                "status": STATUS,
                "timed_out": False,
                "requested_m": requested_m,
                "m": m,
                "machine_count_adjusted": machine_count_adjusted(requested_m, m),
                "c": c,
                "theta": theta,
                "source_scale": inst.scale,
                "cg_objective": cg_obj,
                "gap_vs_cg_pct": gap_vs_cg,
                "wspt_lrf_objective": result.wspt_lrf_objective,
                "batch_reorder_improvement_pct": result.batch_reorder_improvement_pct,
                "heuristic_variant": METHOD_NAME,
            }
        )
    return wspt_rows


def union_header(base_header: list[str], rows: list[dict[str, object]]) -> list[str]:
    header = list(base_header)
    for extra in [
        "m",
        "requested_m",
        "machine_count_adjusted",
        "c",
        "theta",
        "source_scale",
        "cg_objective",
        "gap_vs_cg_pct",
        "wspt_lrf_objective",
        "batch_reorder_improvement_pct",
        "heuristic_variant",
    ]:
        if extra not in header:
            header.append(extra)
    for row in rows:
        for key in row:
            if key not in header:
                header.append(key)
    return header


def write_csv(path: Path, header: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        writer.writerows(rows)


def add_summary_sheet(workbook: Workbook, wspt_rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet("wspt_summary")
    sheet.append(
        [
            "scale",
            "instances",
            "n range",
            "mean gap vs CG (%)",
            "mean WSPT time (s)",
            "mean block-reorder improvement (%)",
        ]
    )
    for scale in ("small", "medium", "large", "all"):
        rows = wspt_rows if scale == "all" else [row for row in wspt_rows if row.get("source_scale") == scale]
        if not rows:
            continue
        ns = [int(row["n"]) for row in rows]
        gaps = [float(row["gap_vs_cg_pct"]) for row in rows if row.get("gap_vs_cg_pct") is not None]
        times = [float(row["time_sec"]) for row in rows]
        improvements = [float(row["batch_reorder_improvement_pct"]) for row in rows]
        sheet.append(
            [
                scale,
                len(rows),
                f"{min(ns)}-{max(ns)}",
                sum(gaps) / len(gaps) if gaps else None,
                sum(times) / len(times) if times else None,
                sum(improvements) / len(improvements) if improvements else None,
            ]
        )


def write_xlsx(path: Path, header: list[str], rows: list[dict[str, object]], wspt_rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "comparison_with_wspt"
    sheet.append(header)
    for row in rows:
        sheet.append([row.get(column) for column in header])
    add_summary_sheet(workbook, wspt_rows)
    workbook.save(path)


def append_wspt_to_saved_results(
    saved_comparison: Path,
    data_dir: Path,
    output_xlsx: Path,
    output_csv: Path | None = None,
    *,
    machines: int | None = None,
) -> list[dict[str, object]]:
    header, source_rows = read_table(saved_comparison)
    instances = load_instances(data_dir)
    wspt_rows = build_wspt_rows(source_rows, instances, machines=machines)
    combined_rows = list(source_rows) + wspt_rows
    combined_header = union_header(header, combined_rows)
    write_xlsx(output_xlsx, combined_header, combined_rows, wspt_rows)
    if output_csv is not None:
        write_csv(output_csv, combined_header, combined_rows)
    return wspt_rows
