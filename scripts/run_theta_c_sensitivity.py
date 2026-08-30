#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CG versus WSPT-BR theta/c sensitivity experiment.

This file is deliberately aligned with the old workbook
``对比实验数据_加入WSPT.xlsx``:

* the 25 job counts are exactly
  6, 8, 10, 12, 14, 16, 18, 20, 25, 30, 40, 50, 60, 70,
  80, 90, 100, 110, 120, 130, 150, 160, 180, 190, 200;
* the main comparison has one instance for every job count and three theta
  levels, giving 75 cases;
* the interaction comparison uses six representatives and c in {2, 4},
  giving 36 more cases;
* every case writes exactly two method rows: CG followed by WSPT-BR;
* ``num_columns`` and ``iterations`` are copied directly from the original
  CG return record. They are never generated or modified by this script.

The default output contains 111 cases and 222 method rows. FullMIP is not part
of the default two-method comparison; use --include-full-mip only when its
additional rows and the CG-FullMIP signed difference are needed.
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from pmcg.parameters import effective_machine_count  # noqa: E402
from pmcg.wspt_baseline import run_improved_wspt  # noqa: E402


# -------------------- Experiment constants --------------------
N_ORDER = (
    6, 8, 10, 12, 14, 16, 18, 20, 25, 30,
    40, 50, 60, 70, 80, 90, 100, 110, 120, 130, 150,
    160, 180, 190, 200,
)
SMALL_N = frozenset({6, 8, 10, 12, 14, 16, 18, 20, 25, 30})
MEDIUM_N = frozenset({40, 50, 60, 70, 80, 90, 100, 110, 120, 150})
LARGE_N = frozenset(set(N_ORDER) - SMALL_N - MEDIUM_N)
SCALE_C = {"small": 4, "medium": 2, "large": 2}
THETA_RATIOS = (0.5, 1.0, 2.0)
INTERACTION_C = (2, 4)
INTERACTION_N = (6, 30, 40, 150, 130, 200)
THETA_LABEL = {0.5: "small", 1.0: "mid", 2.0: "large"}

DEFAULT_M = 3
DEFAULT_K = 30
DEFAULT_TIME_LIMIT = 1800.0
DEFAULT_FULL_MIP_TIME_LIMIT = 1800.0
DEFAULT_SEED = 20260819
DEFAULT_VALUE_MIN = 1
DEFAULT_VALUE_MAX = 36
FINAL_MASTER_RESERVE = 360.0
DEFAULT_OUTPUT = REPO_ROOT / "outputs" / "theta-c-sensitivity" / "theta_c_sensitivity.xlsx"
WSPT_NAME = "WSPT-BR"

# Exact old comparison_with_wspt column order.
OLD_COLUMNS = [
    "instance_id", "n", "method", "objective", "time_sec", "status",
    "timed_out", "bound", "gap", "sol_count", "K", "iterations",
    "num_columns", "initial_columns", "initial_schedules", "pricing_status",
    "pricing_sol_count", "integer_status", "incumbent_status", "objective_source",
    "last_integer_update_iter", "initial_objective", "last_rmp_obj",
    "best_reduced_cost", "final_int_bound", "final_int_gap", "final_int_sol_count",
    "checkpoint_saved_at", "p", "w", "m", "c", "theta", "source_scale",
    "cg_objective", "gap_vs_cg_pct", "wspt_lrf_objective",
    "batch_reorder_improvement_pct", "heuristic_variant",
]
CASE_COLUMNS = [
    "case_id", "experiment", "case_index", "theta_ratio", "theta_label",
    "theta_value", "mean_p", "requested_m", "m_effective", "random_seed",
    "fullmip_objective", "fullmip_status", "cg_minus_fullmip_signed_diff",
    "cg_minus_fullmip_signed_diff_pct",
]
OUTPUT_COLUMNS = OLD_COLUMNS + CASE_COLUMNS


@dataclass(frozen=True)
class Instance:
    n: int
    instance_id: int
    scale: str
    p: list[float]
    w: list[float]


@dataclass(frozen=True)
class Case:
    case_id: str
    experiment: str
    case_index: int
    instance: Instance
    c: int
    theta_ratio: float
    theta: float
    m: int
    m_effective: int


def scale_of_n(n: int) -> str:
    if n in SMALL_N:
        return "small"
    if n in MEDIUM_N:
        return "medium"
    if n in LARGE_N:
        return "large"
    raise ValueError(f"Unexpected n={n}; it is not in the configured instance list.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--m", type=int, default=DEFAULT_M)
    parser.add_argument("--k", type=int, default=DEFAULT_K)
    parser.add_argument("--cg-time-limit", type=float, default=DEFAULT_TIME_LIMIT)
    parser.add_argument("--full-mip-time-limit", type=float, default=DEFAULT_FULL_MIP_TIME_LIMIT)
    parser.add_argument("--include-full-mip", action="store_true")
    parser.add_argument("--random-seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--value-min", type=int, default=DEFAULT_VALUE_MIN)
    parser.add_argument("--value-max", type=int, default=DEFAULT_VALUE_MAX)
    parser.add_argument("--max-cases", type=int, default=0)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Write the case plan and metadata without calling a solver.",
    )
    parser.add_argument(
        "--start-case",
        type=int,
        default=1,
        help="1-based original case index to start from, inclusive.",
    )
    parser.add_argument(
        "--end-case",
        type=int,
        default=0,
        help="1-based original case index to stop at, inclusive. 0 means the last case.",
    )
    args = parser.parse_args()
    if not args.dry_run and args.cg_time_limit <= FINAL_MASTER_RESERVE:
        raise ValueError(
            f"--cg-time-limit must be greater than {FINAL_MASTER_RESERVE:.0f} seconds; "
            "otherwise the original CG reserves all time for final recovery."
        )
    return args


def load_instances(args: argparse.Namespace) -> list[Instance]:
    if args.value_min > args.value_max:
        raise ValueError("value-min must be no greater than value-max")
    source_ids: dict[int, int] = {}
    for scale in ("small", "medium", "large"):
        path = REPO_ROOT / "data" / f"instances_{scale}.json"
        for item in json.loads(path.read_text(encoding="utf-8")):
            source_ids[int(item["n"])] = int(item["id"])

    rng = random.Random(args.random_seed)
    instances: list[Instance] = []
    for n in N_ORDER:
        scale = scale_of_n(n)
        p = [float(rng.randint(args.value_min, args.value_max)) for _ in range(n)]
        w = [float(rng.randint(args.value_min, args.value_max)) for _ in range(n)]
        # New values 18, 20, 25, and 30 have no old JSON record.  Their n is
        # the stable unique id, matching the existing n-as-id convention.
        instances.append(Instance(n, source_ids.get(n, n), scale, p, w))
    return instances


def build_cases(instances: list[Instance], requested_m: int, chosen_n: tuple[int, ...]) -> list[Case]:
    by_n = {instance.n: instance for instance in instances}
    cases: list[Case] = []
    index = 1

    for instance in instances:
        mean_p = float(np.mean(instance.p))
        c = SCALE_C[instance.scale]
        m_eff = effective_machine_count(requested_m, instance.n)
        for ratio in THETA_RATIOS:
            cases.append(
                Case(
                    f"TS-n{instance.n}-r{ratio:.1f}",
                    "theta_sensitivity",
                    index,
                    instance,
                    c,
                    ratio,
                    ratio * mean_p,
                    requested_m,
                    m_eff,
                )
            )
            index += 1

    for n in chosen_n:
        if n not in by_n:
            raise ValueError(f"Interaction n={n} is not in the configured instance list.")
        instance = by_n[n]
        mean_p = float(np.mean(instance.p))
        m_eff = effective_machine_count(requested_m, instance.n)
        for c in INTERACTION_C:
            for ratio in THETA_RATIOS:
                cases.append(
                    Case(
                        f"CI-n{n}-c{c}-r{ratio:.1f}",
                        "theta_c_interaction",
                        index,
                        instance,
                        c,
                        ratio,
                        ratio * mean_p,
                        requested_m,
                        m_eff,
                    )
                )
                index += 1
    return cases


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.integer, np.floating)):
        value = value.item()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return None


def number_list(values: list[float]) -> str:
    return json.dumps(
        [int(value) if float(value).is_integer() else float(value) for value in values],
        separators=(",", ": "),
    )


def empty_or_number(series: pd.Series, mode: str = "mean") -> float | None:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return None
    return float(values.mean() if mode == "mean" else values.median())


def base_row(case: Case, method: str, payload: dict[str, Any], seed: int) -> dict[str, Any]:
    row = {column: None for column in OUTPUT_COLUMNS}
    row.update(
        {
            "instance_id": case.instance.instance_id,
            "n": case.instance.n,
            "method": method,
            "objective": clean(payload.get("objective")),
            "time_sec": clean(payload.get("time_sec")),
            "status": clean(payload.get("status")),
            "timed_out": clean(payload.get("timed_out")),
            "bound": clean(payload.get("bound")),
            "gap": clean(payload.get("gap")),
            "sol_count": clean(payload.get("sol_count")),
            "K": clean(payload.get("K")),
            "iterations": clean(payload.get("iterations")),
            # These are copied from the CG result; no synthetic column count is made.
            "num_columns": clean(payload.get("num_columns")),
            "initial_columns": clean(payload.get("initial_columns")),
            "initial_schedules": clean(payload.get("initial_schedules")),
            "pricing_status": clean(payload.get("pricing_status")),
            "pricing_sol_count": clean(payload.get("pricing_sol_count")),
            "integer_status": clean(payload.get("integer_status")),
            "incumbent_status": clean(payload.get("incumbent_status")),
            "objective_source": clean(payload.get("objective_source")),
            "last_integer_update_iter": clean(payload.get("last_integer_update_iter")),
            "initial_objective": clean(payload.get("initial_objective")),
            "last_rmp_obj": clean(payload.get("last_rmp_obj")),
            "best_reduced_cost": clean(payload.get("best_reduced_cost")),
            "final_int_bound": clean(payload.get("final_int_bound")),
            "final_int_gap": clean(payload.get("final_int_gap")),
            "final_int_sol_count": clean(payload.get("final_int_sol_count")),
            "m": case.m,
            "c": case.c,
            "theta": case.theta,
            "source_scale": case.instance.scale,
            "case_id": case.case_id,
            "experiment": case.experiment,
            "case_index": case.case_index,
            "theta_ratio": case.theta_ratio,
            "theta_label": THETA_LABEL[case.theta_ratio],
            "theta_value": case.theta,
            "mean_p": float(np.mean(case.instance.p)),
            "requested_m": case.m,
            "m_effective": case.m_effective,
            "random_seed": seed,
        }
    )
    # Match the previous workbook: p and w are stored on the solver row.
    if method == "CG":
        row["p"] = number_list(case.instance.p)
        row["w"] = number_list(case.instance.w)
    if method == WSPT_NAME:
        row["wspt_lrf_objective"] = clean(payload.get("wspt_lrf_objective"))
        row["batch_reorder_improvement_pct"] = clean(
            payload.get("batch_reorder_improvement_pct")
        )
        row["heuristic_variant"] = WSPT_NAME
    return row


def run_wspt(case: Case) -> dict[str, Any]:
    result = run_improved_wspt(
        case.instance.p,
        case.instance.w,
        case.m_effective,
        case.c,
        case.theta,
    )
    return {
        "objective": result.objective,
        "time_sec": result.time_sec,
        "status": "FEASIBLE",
        "timed_out": False,
        "wspt_lrf_objective": result.wspt_lrf_objective,
        "batch_reorder_improvement_pct": result.batch_reorder_improvement_pct,
    }


def run_cg(case: Case, args: argparse.Namespace) -> dict[str, Any]:
    # This is the exact repository entry point used by comparison.py.
    from pmcg.ml_assisted import run_fixed_k_column_generation

    return run_fixed_k_column_generation(
        case.instance.p,
        case.instance.w,
        case.m_effective,
        args.k,
        args.cg_time_limit,
        tool_change_interval=case.c,
        tool_change_time=case.theta,
    )


@contextmanager
def temporary_comparison_parameters(case: Case, time_limit: float):
    from pmcg import comparison

    old = {"TIME_LIMIT": comparison.TIME_LIMIT, "c": comparison.c, "theta": comparison.theta}
    comparison.TIME_LIMIT = time_limit
    comparison.c = case.c
    comparison.theta = case.theta
    try:
        yield comparison
    finally:
        comparison.TIME_LIMIT = old["TIME_LIMIT"]
        comparison.c = old["c"]
        comparison.theta = old["theta"]


def run_full_mip(case: Case, args: argparse.Namespace) -> dict[str, Any]:
    with temporary_comparison_parameters(case, args.full_mip_time_limit) as comparison:
        return comparison.solve_full_MIP(case.instance.p, case.instance.w, case.m_effective)


def add_comparison_fields(
    wspt: dict[str, Any],
    cg_objective: float | None,
    full_mip: dict[str, Any] | None,
) -> None:
    wspt_objective = clean(wspt.get("objective"))
    wspt["cg_objective"] = cg_objective
    wspt["gap_vs_cg_pct"] = (
        None
        if wspt_objective is None or cg_objective in (None, 0)
        else 100.0 * (float(wspt_objective) - float(cg_objective)) / float(cg_objective)
    )
    if full_mip is not None:
        full_objective = clean(full_mip.get("objective"))
        wspt["fullmip_objective"] = full_objective
        wspt["fullmip_status"] = clean(full_mip.get("status"))
        if cg_objective is not None and full_objective is not None:
            diff = float(cg_objective) - float(full_objective)
            wspt["cg_minus_fullmip_signed_diff"] = diff
            wspt["cg_minus_fullmip_signed_diff_pct"] = (
                None if full_objective == 0 else 100.0 * diff / float(full_objective)
            )


def create_checkpoint(path: Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(".tmp.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "comparison_with_wspt"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(column) for column in OUTPUT_COLUMNS])

    # Keep the two sensitivity experiments separately visible while retaining
    # the old long-format main table.
    for sheet_name, experiment in (
        ("theta_sensitivity", "theta_sensitivity"),
        ("theta_c_interaction", "theta_c_interaction"),
    ):
        sub = wb.create_sheet(sheet_name)
        sub.append(OUTPUT_COLUMNS)
        for row in rows:
            if row.get("experiment") == experiment:
                sub.append([row.get(column) for column in OUTPUT_COLUMNS])

    instances_sheet = wb.create_sheet("instances")
    instances_sheet.append(["instance_id", "n", "scale", "p", "w", "mean_p", "random_seed"])
    instance_rows: dict[int, dict[str, Any]] = {}
    for row in rows:
        instance_id = row.get("instance_id")
        # WSPT-BR is checkpointed before CG. Prefer the CG row because it
        # carries p and w in the same way as the old comparison workbook.
        if instance_id not in instance_rows or row.get("method") == "CG":
            instance_rows[instance_id] = row
    for instance_id, row in instance_rows.items():
        instances_sheet.append(
            [
                instance_id,
                row.get("n"),
                row.get("source_scale"),
                row.get("p"),
                row.get("w"),
                row.get("mean_p"),
                row.get("random_seed"),
            ]
        )

    summary = wb.create_sheet("case_summary")
    summary_headers = [
        "experiment", "case_id", "n", "scale", "c", "theta_ratio", "theta",
        "CG objective", "WSPT-BR objective", "WSPT-BR gap vs CG (%)",
        "CG time (s)", "WSPT-BR time (s)", "CG iterations", "CG num_columns",
    ]
    summary.append(summary_headers)
    by_case: dict[str, dict[str, Any]] = {}
    for row in rows:
        by_case.setdefault(row["case_id"], {})[row["method"]] = row
    for case_id, methods in by_case.items():
        cg = methods.get("CG", {})
        wspt = methods.get(WSPT_NAME, {})
        source = cg or wspt
        summary.append(
            [
                source.get("experiment"), case_id, source.get("n"), source.get("source_scale"),
                source.get("c"), source.get("theta_ratio"), source.get("theta"),
                cg.get("objective"), wspt.get("objective"), wspt.get("gap_vs_cg_pct"),
                cg.get("time_sec"), wspt.get("time_sec"), cg.get("iterations"),
                cg.get("num_columns"),
            ]
        )

    meta = wb.create_sheet("metadata")
    meta.append(["key", "value"])
    for key, value in metadata.items():
        meta.append([key, value])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for cells in sheet.columns:
            letter = cells[0].column_letter
            width = max((len(str(cell.value)) for cell in cells[:150] if cell.value is not None), default=0)
            sheet.column_dimensions[letter].width = min(width + 2, 32)
    wb.save(temp)
    temp.replace(path)


def metadata(args: argparse.Namespace, cases: list[Case], rows: list[dict[str, Any]], completed: int) -> dict[str, Any]:
    return {
        "format_reference": "对比实验数据_加入WSPT.xlsx / comparison_with_wspt",
        "n_order": ", ".join(map(str, N_ORDER)),
        "n_count": len(N_ORDER),
        "cases_planned": len(cases),
        "cases_completed": completed,
        "rows_written": len(rows),
        "methods_per_case": "CG, WSPT-BR",
        "m": args.m,
        "K": args.k,
        "cg_time_limit": args.cg_time_limit,
        "theta_ratios": "0.5, 1.0, 2.0",
        "interaction_n": ", ".join(map(str, INTERACTION_N)),
        "interaction_c": "2, 4",
        "random_seed": args.random_seed,
        "random_value_range": f"[{args.value_min}, {args.value_max}]",
        "instance_values": "independent random integers in [value-min, value-max]",
        "full_mip_enabled": args.include_full_mip,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def main() -> None:
    args = parse_args()
    instances = load_instances(args)
    all_cases = build_cases(instances, args.m, INTERACTION_N)
    if not 1 <= args.start_case <= len(all_cases):
        raise ValueError(
            f"start-case must be between 1 and {len(all_cases)}, got {args.start_case}"
        )
    end_case = args.end_case or len(all_cases)
    if end_case < args.start_case or end_case > len(all_cases):
        raise ValueError(
            f"end-case must be between start-case and {len(all_cases)}, got {end_case}"
        )
    cases = all_cases[args.start_case - 1 : end_case]
    if args.max_cases > 0:
        cases = cases[: args.max_cases]

    rows: list[dict[str, Any]] = []
    output = args.output
    print(f"n distribution: {N_ORDER}", flush=True)
    print(
        f"Running original cases {args.start_case}-{args.start_case + len(cases) - 1} "
        f"of {len(all_cases)}; selected cases={len(cases)}; output={output}",
        flush=True,
    )
    create_checkpoint(output, rows, metadata(args, all_cases, rows, 0))
    if args.dry_run:
        print("Dry run complete: the 111-case plan was validated; no solver was called.")
        return

    for position, case in enumerate(cases, start=1):
        original_position = case.case_index
        start = time.perf_counter()
        print(
            f"[{original_position}/{len(all_cases)}] {case.case_id}: n={case.instance.n}, "
            f"c={case.c}, theta/mean(p)={case.theta_ratio}",
            flush=True,
        )
        try:
            # The heuristic row is saved immediately, before the CG call.
            wspt_payload = run_wspt(case)
            wspt_row = base_row(case, WSPT_NAME, wspt_payload, args.random_seed)
            rows.append(wspt_row)
            create_checkpoint(output, rows, metadata(args, all_cases, rows, position - 1))

            cg_payload = run_cg(case, args)
            cg_row = base_row(case, "CG", cg_payload, args.random_seed)
            cg_objective = clean(cg_payload.get("objective"))
            full_payload = run_full_mip(case, args) if args.include_full_mip else None
            add_comparison_fields(wspt_row, cg_objective, full_payload)
            if full_payload is not None:
                full_row = base_row(case, "FullMIP", full_payload, args.random_seed)
                add_comparison_fields(full_row, cg_objective, full_payload)
                rows.insert(len(rows) - 1, full_row)
            # Put CG directly before WSPT-BR, as in the old workbook.
            rows.insert(len(rows) - 1, cg_row)
            result_text = (
                f"OK objective={cg_objective}; iterations={cg_payload.get('iterations')}; "
                f"num_columns={cg_payload.get('num_columns')}"
            )
        except Exception as exc:
            error = f"{type(exc).__name__}: {exc}"
            wspt_row["status"] = "ERROR"
            wspt_row["objective_source"] = error[:300]
            error_payload = {
                "status": "ERROR",
                "timed_out": False,
                "objective_source": error[:300],
            }
            rows.insert(len(rows) - 1, base_row(case, "CG", error_payload, args.random_seed))
            result_text = f"ERROR {error}"

        create_checkpoint(output, rows, metadata(args, all_cases, rows, position))
        print(
            f"[{original_position}/{len(all_cases)}] {result_text}; "
            f"elapsed={time.perf_counter() - start:.2f}s; rows={len(rows)}",
            flush=True,
        )

    print(f"Finished. New workbook: {output}")


if __name__ == "__main__":
    main()
