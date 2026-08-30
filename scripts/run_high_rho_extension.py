#!/usr/bin/env python3
"""Extension experiment for the paper's CG and WSPT-BR comparison.

The output follows the long-format workbook used by the previous comparison
experiment.  Every case produces two consecutive rows in the
``comparison_with_wspt`` sheet:

    CG
    WSPT-BR

Extension experiment design:
    * representative instances n = 6, 16, 40, 110, 150, 200;
    * two small, two medium, and two large instances;
    * p_j and w_j are independently generated as integers in [1, 36];
    * c is fixed as in the paper experiment: small c=4, medium/large c=2;
    * theta / mean(p) in {3.0, 4.0};
    * 6 representative instances x 2 theta levels = 12 cases;
    * 24 CG/WSPT-BR rows in total.

The original repository implementation is used directly:
    * pmcg.ml_assisted.run_fixed_k_column_generation
    * pmcg.wspt_baseline.run_improved_wspt

FullMIP is optional because the requested primary comparison is CG versus
WSPT-BR.  Add --include-full-mip when FullMIP rows and CG-FullMIP signed
differences are also required.
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

CURRENT_FILE = Path(__file__).resolve()
REPO_ROOT = CURRENT_FILE.parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from pmcg.parameters import effective_machine_count  # noqa: E402
from pmcg.wspt_baseline import run_improved_wspt  # noqa: E402


DEFAULT_INSTANCE_FILES = {
    "small": REPO_ROOT / "data" / "instances_small.json",
    "medium": REPO_ROOT / "data" / "instances_medium.json",
    "large": REPO_ROOT / "data" / "instances_large.json",
}
DEFAULT_C_BY_SCALE = {"small": 4, "medium": 2, "large": 2}
THETA_RATIOS = (3.0, 4.0)
THETA_LABELS = {3.0: "3x_mean_p", 4.0: "4x_mean_p"}
REPRESENTATIVE_N = (6, 16, 40, 110, 150, 200)
EXTENSION_EXPERIMENT = "theta_extension_3x_4x"
WSPT_METHOD_NAME = "WSPT-BR"
DEFAULT_M = 3
DEFAULT_K = 30
DEFAULT_RANDOM_SEED = 20260819
DEFAULT_VALUE_MIN = 1
DEFAULT_VALUE_MAX = 36
FINAL_MASTER_RESERVE = 360.0
# Match the paper's original comparison experiment.
DEFAULT_CG_TIME_LIMIT = 1800.0
DEFAULT_FULL_MIP_TIME_LIMIT = 1800.0
DEFAULT_OUTPUT = REPO_ROOT / "outputs" / "theta-c-sensitivity" / "high_rho_extension.xlsx"

# These are the exact first 39 columns of the previous comparison workbook.
OLD_COMPARISON_COLUMNS = [
    "instance_id",
    "n",
    "method",
    "objective",
    "time_sec",
    "status",
    "timed_out",
    "bound",
    "gap",
    "sol_count",
    "K",
    "iterations",
    "num_columns",
    "initial_columns",
    "initial_schedules",
    "pricing_status",
    "pricing_sol_count",
    "integer_status",
    "incumbent_status",
    "objective_source",
    "last_integer_update_iter",
    "initial_objective",
    "last_rmp_obj",
    "best_reduced_cost",
    "final_int_bound",
    "final_int_gap",
    "final_int_sol_count",
    "checkpoint_saved_at",
    "p",
    "w",
    "m",
    "c",
    "theta",
    "source_scale",
    "cg_objective",
    "gap_vs_cg_pct",
    "wspt_lrf_objective",
    "batch_reorder_improvement_pct",
    "heuristic_variant",
]

EXTRA_CASE_COLUMNS = [
    "case_id",
    "experiment",
    "case_index",
    "theta_ratio",
    "theta_label",
    "theta_value",
    "mean_p",
    "requested_m",
    "m_effective",
    "fullmip_objective",
    "fullmip_status",
    "cg_minus_fullmip_signed_diff",
    "cg_minus_fullmip_signed_diff_pct",
    "random_seed",
]
OUTPUT_COLUMNS = OLD_COMPARISON_COLUMNS + EXTRA_CASE_COLUMNS


@dataclass(frozen=True)
class InstanceRecord:
    scale: str
    instance_id: int
    n: int
    p: list[float]
    w: list[float]


@dataclass(frozen=True)
class CaseConfig:
    experiment: str
    case_id: str
    case_index: int
    scale: str
    instance_id: int
    n: int
    requested_m: int
    m_effective: int
    c_value: int
    theta_ratio: float
    theta_label: str
    theta_value: float
    mean_p: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--m", type=int, default=DEFAULT_M)
    parser.add_argument("--k", type=int, default=DEFAULT_K)
    parser.add_argument("--cg-time-limit", type=float, default=DEFAULT_CG_TIME_LIMIT)
    parser.add_argument("--full-mip-time-limit", type=float, default=DEFAULT_FULL_MIP_TIME_LIMIT)
    parser.add_argument("--include-full-mip", action="store_true")
    parser.add_argument("--random-seed", type=int, default=DEFAULT_RANDOM_SEED)
    parser.add_argument("--value-min", type=int, default=DEFAULT_VALUE_MIN)
    parser.add_argument("--value-max", type=int, default=DEFAULT_VALUE_MAX)
    parser.add_argument(
        "--use-file-values",
        action="store_true",
        help="Use the JSON p and w values instead of random integers in [value-min, value-max].",
    )
    parser.add_argument("--max-cases", type=int, default=0)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Write the 12-case plan and metadata without calling a solver.",
    )
    args = parser.parse_args()
    # The original CG reserves FINAL_MASTER_RESERVE = 360 seconds for final
    # integer recovery. A total time limit at or below that reserve skips the
    # regular column-generation loop and is not a valid CG comparison run.
    if not args.dry_run and args.cg_time_limit <= FINAL_MASTER_RESERVE:
        raise ValueError(
            f"--cg-time-limit must be greater than {FINAL_MASTER_RESERVE:.0f} seconds. "
            "The original CG reserves that amount for final integer recovery; "
            "a smaller total limit would skip normal column-generation pricing."
        )
    return args


def load_instances(
    random_seed: int,
    value_min: int,
    value_max: int,
    use_file_values: bool,
) -> list[InstanceRecord]:
    if value_min > value_max:
        raise ValueError("value-min must be no greater than value-max")
    rng = random.Random(random_seed)
    records: list[InstanceRecord] = []
    for scale in ("small", "medium", "large"):
        path = DEFAULT_INSTANCE_FILES[scale]
        for item in json.loads(path.read_text(encoding="utf-8")):
            n = int(item["n"])
            if use_file_values:
                p = [float(value) for value in item["p"]]
                w = [float(value) for value in item["w"]]
            else:
                p = [float(rng.randint(value_min, value_max)) for _ in range(n)]
                w = [float(rng.randint(value_min, value_max)) for _ in range(n)]
            if len(p) != n or len(w) != n:
                raise ValueError(f"Invalid p/w length for instance {item['id']}")
            records.append(InstanceRecord(scale, int(item["id"]), n, p, w))
    scale_order = {"small": 0, "medium": 1, "large": 2}
    return sorted(records, key=lambda x: (scale_order[x.scale], x.n, x.instance_id))


def build_case_plan(
    instances: list[InstanceRecord],
    requested_m: int,
) -> list[CaseConfig]:
    by_id = {instance.instance_id: instance for instance in instances}
    by_n = {instance.n: instance for instance in instances}
    selected_ids = {by_n[n].instance_id for n in REPRESENTATIVE_N}
    missing = selected_ids.difference(by_id)
    if missing:
        raise ValueError(f"Unknown representative instance ids: {sorted(missing)}")

    cases: list[CaseConfig] = []
    case_index = 1

    def append_case(instance: InstanceRecord, experiment: str, c_value: int) -> None:
        nonlocal case_index
        mean_p = float(np.mean(instance.p))
        m_effective = effective_machine_count(requested_m, instance.n)
        for ratio in THETA_RATIOS:
            cases.append(
                CaseConfig(
                    experiment=experiment,
                    case_id=(
                        f"EXT-n{instance.n}-r{ratio:.1f}"
                    ),
                    case_index=case_index,
                    scale=instance.scale,
                    instance_id=instance.instance_id,
                    n=instance.n,
                    requested_m=requested_m,
                    m_effective=m_effective,
                    c_value=c_value,
                    theta_ratio=ratio,
                    theta_label=THETA_LABELS[ratio],
                    theta_value=mean_p * ratio,
                    mean_p=mean_p,
                )
            )
            case_index += 1

    selected_instances = [
        by_id[by_n[n].instance_id]
        for n in REPRESENTATIVE_N
        if by_n[n].instance_id in selected_ids
    ]
    for instance in selected_instances:
        # c remains fixed so that this extension isolates the effect of a
        # larger tool-change duration rather than mixing theta and c effects.
        append_case(instance, EXTENSION_EXPERIMENT, DEFAULT_C_BY_SCALE[instance.scale])
    if len(cases) != 12:
        raise ValueError(f"Expected 12 extension cases, got {len(cases)}")
    return cases


def scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.integer, np.floating)):
        value = value.item()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return None


def json_vector(values: list[float]) -> str:
    return json.dumps(
        [int(value) if float(value).is_integer() else float(value) for value in values],
        ensure_ascii=False,
        separators=(",", ": "),
    )


def payload_to_old_columns(
    case: CaseConfig,
    instance: InstanceRecord,
    method: str,
    payload: dict[str, Any],
    cg_objective: float | None,
    fullmip: dict[str, Any] | None,
    random_seed: int,
) -> dict[str, Any]:
    row = {column: None for column in OUTPUT_COLUMNS}
    row.update(
        {
            "instance_id": case.instance_id,
            "n": case.n,
            "method": method,
            "objective": scalar(payload.get("objective")),
            "time_sec": scalar(payload.get("time_sec")),
            "status": scalar(payload.get("status")),
            "timed_out": scalar(payload.get("timed_out")),
            "bound": scalar(payload.get("bound")),
            "gap": scalar(payload.get("gap")),
            "sol_count": scalar(payload.get("sol_count")),
            "K": scalar(payload.get("K")),
            "iterations": scalar(payload.get("iterations")),
            "num_columns": scalar(payload.get("num_columns")),
            "initial_columns": scalar(payload.get("initial_columns")),
            "initial_schedules": scalar(payload.get("initial_schedules")),
            "pricing_status": scalar(payload.get("pricing_status")),
            "pricing_sol_count": scalar(payload.get("pricing_sol_count")),
            "integer_status": scalar(payload.get("integer_status")),
            "incumbent_status": scalar(payload.get("incumbent_status")),
            "objective_source": scalar(payload.get("objective_source")),
            "last_integer_update_iter": scalar(payload.get("last_integer_update_iter")),
            "initial_objective": scalar(payload.get("initial_objective")),
            "last_rmp_obj": scalar(payload.get("last_rmp_obj")),
            "best_reduced_cost": scalar(payload.get("best_reduced_cost")),
            "final_int_bound": scalar(payload.get("final_int_bound")),
            "final_int_gap": scalar(payload.get("final_int_gap")),
            "final_int_sol_count": scalar(payload.get("final_int_sol_count")),
            "p": json_vector(instance.p) if method == "CG" else None,
            "w": json_vector(instance.w) if method == "CG" else None,
            "m": case.requested_m,
            "c": case.c_value,
            "theta": case.theta_value,
            "source_scale": instance.scale,
            "case_id": case.case_id,
            "experiment": case.experiment,
            "case_index": case.case_index,
            "theta_ratio": case.theta_ratio,
            "theta_label": case.theta_label,
            "theta_value": case.theta_value,
            "mean_p": case.mean_p,
            "requested_m": case.requested_m,
            "m_effective": case.m_effective,
            "random_seed": random_seed,
        }
    )

    if method == WSPT_METHOD_NAME:
        row["cg_objective"] = cg_objective
        wspt_objective = row["objective"]
        row["gap_vs_cg_pct"] = (
            None
            if wspt_objective is None or cg_objective in (None, 0)
            else 100.0 * (float(wspt_objective) - float(cg_objective)) / float(cg_objective)
        )
        row["wspt_lrf_objective"] = scalar(payload.get("wspt_lrf_objective"))
        row["batch_reorder_improvement_pct"] = scalar(
            payload.get("batch_reorder_improvement_pct")
        )
        row["heuristic_variant"] = WSPT_METHOD_NAME

    if fullmip is not None:
        full_objective = scalar(fullmip.get("objective"))
        cg_value = cg_objective
        row["fullmip_objective"] = full_objective
        row["fullmip_status"] = scalar(fullmip.get("status"))
        signed_diff = (
            None if cg_value is None or full_objective is None else float(cg_value) - float(full_objective)
        )
        signed_diff_pct = (
            None
            if signed_diff is None or full_objective in (None, 0)
            else 100.0 * signed_diff / float(full_objective)
        )
        row["cg_minus_fullmip_signed_diff"] = signed_diff
        row["cg_minus_fullmip_signed_diff_pct"] = signed_diff_pct

    return row


def run_wspt_case(instance: InstanceRecord, case: CaseConfig) -> dict[str, Any]:
    result = run_improved_wspt(
        instance.p,
        instance.w,
        case.m_effective,
        case.c_value,
        case.theta_value,
    )
    return {
        "objective": result.objective,
        "time_sec": result.time_sec,
        "status": "FEASIBLE",
        "timed_out": False,
        "wspt_lrf_objective": result.wspt_lrf_objective,
        "batch_reorder_improvement_pct": result.batch_reorder_improvement_pct,
    }


def run_cg_case(instance: InstanceRecord, case: CaseConfig, k: int, time_limit: float) -> dict[str, Any]:
    from pmcg.ml_assisted import run_fixed_k_column_generation

    return run_fixed_k_column_generation(
        instance.p,
        instance.w,
        case.m_effective,
        k,
        time_limit,
        tool_change_interval=case.c_value,
        tool_change_time=case.theta_value,
    )


@contextmanager
def temporary_module_attrs(module: Any, **attrs: Any):
    old = {name: getattr(module, name) for name in attrs}
    for name, value in attrs.items():
        setattr(module, name, value)
    try:
        yield
    finally:
        for name, value in old.items():
            setattr(module, name, value)


def run_fullmip_case(
    instance: InstanceRecord,
    case: CaseConfig,
    time_limit: float,
) -> dict[str, Any]:
    from pmcg import comparison

    with temporary_module_attrs(
        comparison,
        TIME_LIMIT=float(time_limit),
        c=int(case.c_value),
        theta=float(case.theta_value),
    ):
        return comparison.solve_full_MIP(instance.p, instance.w, case.m_effective)


def rows_for_case(
    instance: InstanceRecord,
    case: CaseConfig,
    args: argparse.Namespace,
    random_seed: int,
) -> list[dict[str, Any]]:
    cg_payload = run_cg_case(instance, case, args.k, args.cg_time_limit)
    wspt_payload = run_wspt_case(instance, case)
    fullmip_payload = (
        run_fullmip_case(instance, case, args.full_mip_time_limit)
        if args.include_full_mip
        else None
    )
    cg_objective = scalar(cg_payload.get("objective"))

    cg_row = payload_to_old_columns(
        case, instance, "CG", cg_payload, cg_objective, fullmip_payload, random_seed
    )
    wspt_row = payload_to_old_columns(
        case,
        instance,
        WSPT_METHOD_NAME,
        wspt_payload,
        cg_objective,
        fullmip_payload,
        random_seed,
    )
    # The old workbook appends the WSPT row after the corresponding CG row.
    rows = [cg_row, wspt_row]
    if fullmip_payload is not None:
        fullmip_row = payload_to_old_columns(
            case,
            instance,
            "FullMIP",
            fullmip_payload,
            cg_objective,
            fullmip_payload,
            random_seed,
        )
        rows.insert(0, fullmip_row)
    return rows


def make_wspt_row(
    instance: InstanceRecord,
    case: CaseConfig,
    args: argparse.Namespace,
    random_seed: int,
) -> dict[str, Any]:
    wspt_payload = run_wspt_case(instance, case)
    return payload_to_old_columns(
        case,
        instance,
        WSPT_METHOD_NAME,
        wspt_payload,
        None,
        None,
        random_seed,
    )


def make_cg_row(
    instance: InstanceRecord,
    case: CaseConfig,
    args: argparse.Namespace,
    random_seed: int,
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    cg_payload = run_cg_case(instance, case, args.k, args.cg_time_limit)
    cg_objective = scalar(cg_payload.get("objective"))
    fullmip_payload = (
        run_fullmip_case(instance, case, args.full_mip_time_limit)
        if args.include_full_mip
        else None
    )
    cg_row = payload_to_old_columns(
        case,
        instance,
        "CG",
        cg_payload,
        cg_objective,
        fullmip_payload,
        random_seed,
    )
    return cg_row, fullmip_payload


def complete_wspt_row(
    wspt_row: dict[str, Any],
    instance: InstanceRecord,
    case: CaseConfig,
    cg_objective: float | None,
    fullmip_payload: dict[str, Any] | None,
    random_seed: int,
) -> dict[str, Any]:
    wspt_payload = {
        "objective": wspt_row.get("objective"),
        "time_sec": wspt_row.get("time_sec"),
        "status": wspt_row.get("status"),
        "timed_out": wspt_row.get("timed_out"),
        "wspt_lrf_objective": wspt_row.get("wspt_lrf_objective"),
        "batch_reorder_improvement_pct": wspt_row.get(
            "batch_reorder_improvement_pct"
        ),
    }
    return payload_to_old_columns(
        case,
        instance,
        WSPT_METHOD_NAME,
        wspt_payload,
        cg_objective,
        fullmip_payload,
        random_seed,
    )


def add_summary_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    frame = pd.DataFrame(rows)
    sheet = workbook.create_sheet("wspt_summary")
    headers = [
        "experiment",
        "scale",
        "theta_ratio",
        "c",
        "cases",
        "mean gap vs CG (%)",
        "median gap vs CG (%)",
        "mean WSPT time (s)",
        "mean CG time (s)",
        "mean WSPT objective",
        "mean CG objective",
    ]
    sheet.append(headers)
    if frame.empty:
        return
    wspt = frame[frame["method"] == WSPT_METHOD_NAME].copy()
    # A checkpoint can be written after WSPT-BR but before its CG run finishes.
    # Remove placeholder WSPT columns before merging the available CG rows.
    wspt = wspt.drop(columns=["cg_objective"], errors="ignore")
    cg = frame[frame["method"] == "CG"][["case_id", "time_sec", "objective"]].rename(
        columns={"time_sec": "cg_time_sec", "objective": "cg_objective"}
    )
    wspt = wspt.merge(cg, on="case_id", how="left")
    groups = wspt.groupby(["experiment", "source_scale", "theta_ratio", "c"], dropna=False)
    for (experiment, scale, ratio, c_value), group in groups:
        sheet.append(
            [
                experiment,
                scale,
                ratio,
                c_value,
                int(len(group)),
                scalar(group["gap_vs_cg_pct"].mean()),
                scalar(group["gap_vs_cg_pct"].median()),
                scalar(group["time_sec"].mean()),
                scalar(group["cg_time_sec"].mean()),
                scalar(group["objective"].mean()),
                scalar(group["cg_objective"].mean()),
            ]
        )


def write_workbook(path: Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(".tmp.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "comparison_with_wspt"
    sheet.append(OUTPUT_COLUMNS)
    for row in rows:
        sheet.append([row.get(column) for column in OUTPUT_COLUMNS])
    add_summary_sheet(workbook, rows)

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata.items():
        metadata_sheet.append([key, value])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            max_length = max(
                (len(str(cell.value)) for cell in column_cells[:200] if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 32)
    workbook.save(temp_path)
    temp_path.replace(path)


def build_metadata(
    args: argparse.Namespace,
    cases_total: int,
    rows: list[dict[str, Any]],
    completed_cases: int,
) -> dict[str, Any]:
    return {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "repo_root": str(REPO_ROOT),
        "script_path": str(CURRENT_FILE),
        "cases_planned": cases_total,
        "cases_completed": completed_cases,
        "cg_rows": sum(row.get("method") == "CG" for row in rows),
        "wspt_br_rows": sum(row.get("method") == WSPT_METHOD_NAME for row in rows),
        "fullmip_enabled": args.include_full_mip,
        "random_values": not args.use_file_values,
        "random_seed": args.random_seed,
        "random_value_range": f"[{args.value_min}, {args.value_max}]",
        "experiment": EXTENSION_EXPERIMENT,
        "representative_n": ", ".join(map(str, REPRESENTATIVE_N)),
        "theta_ratios": "3.0, 4.0",
        "c_by_scale": "small=4; medium=2; large=2",
        "m": args.m,
        "K": args.k,
        "cg_time_limit_sec": args.cg_time_limit,
        "full_mip_time_limit_sec": args.full_mip_time_limit if args.include_full_mip else None,
        "format_reference": "对比实验数据_加入WSPT.xlsx / comparison_with_wspt",
    }


def error_rows_for_case(
    instance: InstanceRecord,
    case: CaseConfig,
    error: str,
    random_seed: int,
) -> list[dict[str, Any]]:
    error_payload = {
        "objective": None,
        "time_sec": None,
        "status": "ERROR",
        "timed_out": False,
        "error": error,
    }
    rows = [
        payload_to_old_columns(
            case, instance, "CG", error_payload, None, None, random_seed
        ),
        payload_to_old_columns(
            case, instance, WSPT_METHOD_NAME, error_payload, None, None, random_seed
        ),
    ]
    for row in rows:
        row["status"] = "ERROR"
        row["objective_source"] = error[:300]
    return rows


def main() -> None:
    args = parse_args()
    instances = load_instances(
        random_seed=args.random_seed,
        value_min=args.value_min,
        value_max=args.value_max,
        use_file_values=args.use_file_values,
    )
    cases = build_case_plan(instances, args.m)
    if args.max_cases > 0:
        cases = cases[: args.max_cases]
    instance_by_id = {instance.instance_id: instance for instance in instances}

    rows: list[dict[str, Any]] = []
    output_path = args.output
    print(
        f"Planned {len(cases)} cases; each case writes CG and WSPT-BR rows.",
        flush=True,
    )
    print(
        f"CG time limit per case: {args.cg_time_limit:.1f} s; output: {output_path}",
        flush=True,
    )

    # Create the workbook immediately, before the first potentially long CG run.
    write_workbook(
        output_path,
        rows,
        build_metadata(args, len(cases), rows, completed_cases=0),
    )
    print(f"Initial workbook created: {output_path}", flush=True)
    if args.dry_run:
        print("Dry run complete: the 12-case plan was validated; no solver was called.")
        return

    for completed_index, case in enumerate(cases, start=1):
        instance = instance_by_id[case.instance_id]
        started = time.perf_counter()
        print(
            f"[{completed_index}/{len(cases)}] START {case.case_id} "
            f"scale={case.scale} n={case.n} c={case.c_value} "
            f"theta/mean(p)={case.theta_ratio:.1f}",
            flush=True,
        )
        try:
            # WSPT-BR is solver-free, so save it before the potentially long CG run.
            provisional_wspt = make_wspt_row(instance, case, args, args.random_seed)
            rows.append(provisional_wspt)
            write_workbook(
                output_path,
                rows,
                build_metadata(args, len(cases), rows, completed_index - 1),
            )
            print(
                f"[{completed_index}/{len(cases)}] WSPT-BR saved; "
                f"starting CG; checkpoint rows={len(rows)}",
                flush=True,
            )

            cg_row, fullmip_payload = make_cg_row(
                instance, case, args, args.random_seed
            )
            cg_objective = scalar(cg_row.get("objective"))
            completed_wspt = complete_wspt_row(
                provisional_wspt,
                instance,
                case,
                cg_objective,
                fullmip_payload,
                args.random_seed,
            )
            rows[-1] = completed_wspt
            # Match the old workbook convention: CG immediately before WSPT-BR.
            rows.insert(len(rows) - 1, cg_row)
            if fullmip_payload is not None:
                fullmip_row = payload_to_old_columns(
                    case,
                    instance,
                    "FullMIP",
                    fullmip_payload,
                    cg_objective,
                    fullmip_payload,
                    args.random_seed,
                )
                rows.insert(len(rows) - 2, fullmip_row)
            status = "OK"
        except Exception as exc:  # keep earlier cases in the checkpoint workbook
            error = f"{type(exc).__name__}: {exc}"
            if rows and rows[-1].get("case_id") == case.case_id:
                rows[-1]["status"] = "ERROR"
                rows[-1]["objective_source"] = error[:300]
            else:
                rows.extend(error_rows_for_case(instance, case, error, args.random_seed))
            status = f"ERROR ({error})"

        elapsed = time.perf_counter() - started
        metadata = build_metadata(args, len(cases), rows, completed_index)
        write_workbook(output_path, rows, metadata)
        print(
            f"[{completed_index}/{len(cases)}] {status}; "
            f"elapsed={elapsed:.2f}s; checkpoint rows={len(rows)}; "
            f"saved={output_path}",
            flush=True,
        )

    print(f"Finished. Cases: {len(cases)}; rows: {len(rows)}; output: {output_path}")


if __name__ == "__main__":
    main()
